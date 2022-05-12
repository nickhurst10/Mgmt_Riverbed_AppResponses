"""
This is a script to get configuration data from multiple Riverbed AppResponse 11 (AR11) and present them in a spreadsheet. The script requires two other libraries to work, “appresponse_device.py” and “appresponse_mgmt_api.py”, both of which are part of the github repository “https://github.com/nickhurst10/Manage_Riverbed_AppResponses.git”.
“appresponse_mgmt_api.py” interacts with the AR11 to do the restAPI GET requests.
“appresponse_device.py” manages the interaction between the main script and “appresponse_mgmt_api.py” library.

For the script to know what AR11 to access, this script looks for a file called “ar_list.csv” and looks for IP addresses under the header “ar_list”. An example is include in the github repository.

To run the script, the user must provide user credentials, which have the relevant access to all the AR11’s.

To this script was tested on python version 3.9.

Example to run the script
	python3 Manage_ARs_RO.py -u username

After which the user will be prompted to enter their password.

If you have any questions, please reach out to me.
"""
__author__ = "Nick Hurst nhurst@riverbed.com"

import requests
import json
from getpass import getpass
from datetime import datetime
import openpyxl
import appresponse_device
import argparse
import csv



COLUMN_TITLE_NUMBER = 1
MGMT_SPREADSHEET_PATH = 'AR_Management.xlsx'
AR_LIST_COL_NAME = 'AR_IP'
API_CONFIG_FILE_NAME = 'api_config_info.conf'
AR_LIST_CSV_FILE_PATH = 'ar_list.csv'


AR_LIST_WORKSHEET  = 'AR_List'
CONFIG_OPTIONS_LIST = ["snmp","dns","ntp","hostgroups","urls","apps"]
NTP_WORKSHEET_NAME = 'ntp_information'
SNMP_WORKSHEET_NAME = 'snmp_info'
COMMOM_WORKSHEET_NAME = "common_info"
VIFGS_WORKSHEET_NAME = "vifgs_info"
CAP_JOBS_WORKSHEET_NAME = "cap_jobs_info"
DNS_WORKSHEET_NAME = "dns_info"
PHY_INT_WORKSHEET_NAME = "phy_int_info"
HOSTGROUPS_WORKSHEET_NAME = "hostgroup_info"
URL_WORKSHEET_NAME = "url_info"
APPLICATIONS_WORKSHEET_NAME = "apps_info"


NTP_COL_NAME_LIST = ["server_id","address","prefer","version","encryption","key_id","secret"]
SNMP_COL_NAME_LIST = ["enabled","contact","location","description","version","community_string","username","security_model","auth_protocol","authentication_passphrase","privacy_protocol","privacy_passphrase"]
COMMON_COL_NAME_LIST = ["device_name","sw_version","hw_version","mgmt_addresses","serial","model"]
VIFGS_COL_NAME_LIST = ["name","id","enabled","dedup","interfaces","description","filer_type","filter_value","bandwidth_capacity","is_other_vifg"]
CAP_JOBS_COL_NAME_LIST = ["name","vifgs","capture_from_all_vifgs","snap_len","enabled","microflow_index_enable","min_disk_space","max_disk_space","min_retention_time","max_retention_time","packet_data_optimize_for_read","min_disk_space","max_disk_space","min_retention_time","max_retention_time"]
DNS_COL_NAME_LIST = ["hostname","dns_servers","dns_domains"]
PHY_INT_COL_NAME_LIST = ["name","enabled","description","speed_duplex","interface_type","status","mtu"]
HOSTGROUPS_COL_NAME_LIST = ["name","description","enabled","hosts","id (RO)","member_hostgroups","in_speed","in_speed_unit","out_speed","out_speed_unit"]
URL_COL_NAME_LIST = ["name","description","enabled","preferred","urls","id"]
APPLICATIONS_COL_NAME_LIST = ["name","desc","enable","traffic_match_mode","include_dpi_tags","hosts","ports","ip_protocol","id"]


def string_to_list(received_string):
    disallowed_characters = ["[","]","'"," "]
    return_list = []
    for disallowed_character in disallowed_characters:
        received_string = received_string.replace(disallowed_character,"")
    if received_string == "":
        return_list== []
    else:
        return_list = received_string.split(",")
    return return_list

def string_of_strings_to_list(received_string):
    disallowed_characters = ["[","]","'"]
    return_list = []
    for disallowed_character in disallowed_characters:
        received_string = received_string.replace(disallowed_character,"")
    if received_string == "":
        return_list== []
    else:
        return_list = received_string.split(", ") #had to split on ", " not just "," due to space in list to string form AR11 into excel
    return return_list
    
def setup_worksheet(received_work_book,received_worksheet_name,received_columns_names_list):
    if received_worksheet_name in received_work_book.sheetnames:
        worksheet = received_work_book[received_worksheet_name]
        received_work_book.remove(worksheet)
    
    received_work_book.create_sheet(received_worksheet_name)
    worksheet = received_work_book[received_worksheet_name]

    worksheet.cell(1,1).value = AR_LIST_COL_NAME
    col_num = 2
    for col_name in received_columns_names_list:
        worksheet.cell(1,col_num).value = col_name
        col_num += 1

def setup_management_spreadsheet(received_spreadsheet_path):

    ar_config_work_book = openpyxl.Workbook()

    setup_worksheet(ar_config_work_book,NTP_WORKSHEET_NAME,NTP_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,SNMP_WORKSHEET_NAME,SNMP_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,COMMOM_WORKSHEET_NAME,COMMON_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,VIFGS_WORKSHEET_NAME,VIFGS_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,CAP_JOBS_WORKSHEET_NAME,CAP_JOBS_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,DNS_WORKSHEET_NAME,DNS_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,PHY_INT_WORKSHEET_NAME,PHY_INT_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,HOSTGROUPS_WORKSHEET_NAME,HOSTGROUPS_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,APPLICATIONS_WORKSHEET_NAME,APPLICATIONS_COL_NAME_LIST)
    setup_worksheet(ar_config_work_book,URL_WORKSHEET_NAME,URL_COL_NAME_LIST)

    ar_config_work_book.save(received_spreadsheet_path)
    
def update_spreadsheet_with_config(received_spreadsheet_path,received_ar_config):

    #>>> source = wb.active
    #>>> target = wb.copy_worksheet(source)

    ar_config_work_book = openpyxl.load_workbook(received_spreadsheet_path)
    #fill all relevant worksheets with the there relevant AR information
    update_ntp_worksheet(received_ar_config.ip_addr,received_ar_config.ntp_original_data,ar_config_work_book,NTP_WORKSHEET_NAME)
    update_snmp_worksheet(received_ar_config.ip_addr,received_ar_config.snmp_original_data,ar_config_work_book,SNMP_WORKSHEET_NAME)
    update_vifgs_worksheet(received_ar_config.ip_addr,received_ar_config.vifgs_original_data,ar_config_work_book,VIFGS_WORKSHEET_NAME)
    update_common_worksheet(received_ar_config.ip_addr,received_ar_config.common_original_data,ar_config_work_book,COMMOM_WORKSHEET_NAME)
    update_dns_worksheet(received_ar_config.ip_addr,received_ar_config.dns_original_data,ar_config_work_book,DNS_WORKSHEET_NAME)
    update_cap_jobs_worksheet(received_ar_config.ip_addr,received_ar_config.cap_job_original_data,ar_config_work_book,CAP_JOBS_WORKSHEET_NAME)
    update_phy_int_worksheet(received_ar_config.ip_addr,received_ar_config.phy_int_original_data,ar_config_work_book,PHY_INT_WORKSHEET_NAME)
    update_apps_worksheet(received_ar_config.ip_addr,received_ar_config.ar_apps_original_data,ar_config_work_book,APPLICATIONS_WORKSHEET_NAME)
    update_hostgroups_worksheet(received_ar_config.ip_addr,received_ar_config.hostgroups_original_data,ar_config_work_book,HOSTGROUPS_WORKSHEET_NAME)
    update_urls_worksheet(received_ar_config.ip_addr,received_ar_config.urls_original_data,ar_config_work_book,URL_WORKSHEET_NAME)
    ar_config_work_book.save(received_spreadsheet_path)

def find_column_with_title(received_title_name,received_worksheet):
    for col_num in range(COLUMN_TITLE_NUMBER,received_worksheet.max_column+1):
        if received_worksheet.cell(1,col_num).value == received_title_name:
            return col_num

def get_ar_bearer_token(received_username,received_password,received_ar_ip_addr):
    #will either return bearer token or a blank string to mean rest api access wasn't successful
    bearer_token=""
    url = f"https://{received_ar_ip_addr}/api/mgmt.aaa/1.0/token"
    payload = json.dumps({
        "user_credentials": {
            "username": received_username,
            "password": received_password
        }
        })
    headers = {
        'Content-Type': 'application/json'
        }
    try:
        response = requests.request("POST", url, headers=headers, data=payload,verify=False)
    except Exception as error_message:
        print (error_message)
    else:
        try:
            json_response = json.loads(response.text)
            print(f"bearer Token is: {json_response['access_token']}")
            bearer_token = json_response['access_token']
        except:
            print(f'failed to get bearer token from AR {received_ar_ip_addr} - data received was - \n\t{response.text}')
    return bearer_token

def confirm_rest_api_access_to_ARs_and_get_bearer_token(received_username,received_password,received_ar_ip_addr_list):
    ar_ip_addr_and_bearer_token_list = []
    for ar_ip_addr in received_ar_ip_addr_list:
        print(ar_ip_addr)
        bearer_token = (get_ar_bearer_token(received_username,received_password,ar_ip_addr))
        #if the received bearer token isn't blank then we know we can access the REST API of the AR so we add to the AR list with bearer token
        if bearer_token !="":
            ar_ip_addr_and_bearer_token={'bear_token':bearer_token, 'ar_ip_addr':ar_ip_addr}
            ar_ip_addr_and_bearer_token_list.append(ar_ip_addr_and_bearer_token)
    return (ar_ip_addr_and_bearer_token_list)

def update_ntp_worksheet(received_ar_ip_addr,received_ntp_config_list,received_workbook,received_ntp_worksheet_name):
    
    ntp_work_sheet = received_workbook[received_ntp_worksheet_name]
       
    for ntp_data in received_ntp_config_list['items']:
        row_num=ntp_work_sheet.max_row+1
        ntp_work_sheet.cell(row_num,1).value=received_ar_ip_addr
        ntp_work_sheet.cell(row_num,2).value=ntp_data['server_id']
        ntp_work_sheet.cell(row_num,3).value=ntp_data['address']
        ntp_work_sheet.cell(row_num,4).value=ntp_data['prefer']
        ntp_work_sheet.cell(row_num,5).value=ntp_data['version']
        ntp_work_sheet.cell(row_num,6).value=ntp_data['encryption']
        #if there is encryption add to worksheet
        if (ntp_work_sheet.cell(row_num,6).value == 'md5' or 
            ntp_work_sheet.cell(row_num,6).value =='sha1'):

            ntp_work_sheet.cell(row_num,7).value=ntp_data['key_id']
            ntp_work_sheet.cell(row_num,8).value=ntp_data['secret']

def update_snmp_worksheet(received_ar_ip_addr,received_snmp_config,received_workbook,received_snmp_worksheet_name):
    
    snmp_work_sheet = received_workbook[received_snmp_worksheet_name]
    row_num = snmp_work_sheet.max_row+1
    snmp_work_sheet.cell(row_num,1).value=received_ar_ip_addr
    snmp_work_sheet.cell(row_num,2).value=received_snmp_config.get('enabled')
    snmp_work_sheet.cell(row_num,3).value=received_snmp_config.get('contact')
    snmp_work_sheet.cell(row_num,4).value=received_snmp_config.get('location')
    snmp_work_sheet.cell(row_num,5).value=received_snmp_config.get('description')
    version_configuration = received_snmp_config['version_configuration']
    snmp_work_sheet.cell(row_num,6).value=version_configuration.get('version')
    snmp_work_sheet.cell(row_num,7).value=version_configuration.get('community_string')
    snmp_work_sheet.cell(row_num,8).value=version_configuration.get('username')
    snmp_work_sheet.cell(row_num,9).value=version_configuration.get('security_model')
    snmp_work_sheet.cell(row_num,10).value=version_configuration.get('auth_protocol')
    snmp_work_sheet.cell(row_num,11).value=version_configuration.get('authentication_passphrase')
    snmp_work_sheet.cell(row_num,12).value=version_configuration.get('privacy_protocol')
    snmp_work_sheet.cell(row_num,13).value=version_configuration.get('privacy_passphrase')

def update_vifgs_worksheet(received_ar_ip_addr,received_vifgs_configs,received_workbook,received_vifgs_worksheet_name):
    
    vifgs_work_sheet = received_workbook[received_vifgs_worksheet_name]

    for vifgs_data in received_vifgs_configs["items"]:
        row_num = vifgs_work_sheet.max_row+1
        vifgs_config = vifgs_data["config"]
        vifgs_work_sheet.cell(row_num,1).value=received_ar_ip_addr
        vifgs_work_sheet.cell(row_num,2).value=vifgs_config["name"]
        vifgs_work_sheet.cell(row_num,3).value=vifgs_data["id"]
        vifgs_work_sheet.cell(row_num,4).value=vifgs_config["enabled"]
        vifgs_work_sheet.cell(row_num,5).value=vifgs_config["dedup"]
        vifgs_work_sheet.cell(row_num,6).value=str(vifgs_config["members"])
        vifgs_work_sheet.cell(row_num,7).value=vifgs_config["description"]
        vifgs_work_sheet.cell(row_num,8).value=vifgs_config["filter"]["type"]
        vifgs_work_sheet.cell(row_num,9).value=vifgs_config["filter"]["value"]
        vifgs_work_sheet.cell(row_num,10).value=vifgs_config["bandwidth_capacity"]
        vifgs_work_sheet.cell(row_num,11).value=vifgs_config["is_other_vifg"]

def update_common_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):

    common_worksheet = received_workbook[received_worksheet_name]
    row_num = common_worksheet.max_row+1
    common_worksheet.cell(row_num,1).value=received_ar_ip_addr
    common_worksheet.cell(row_num,2).value=received_configs.get("device_name")
    common_worksheet.cell(row_num,3).value=received_configs.get("sw_version")
    common_worksheet.cell(row_num,4).value=received_configs.get("hw_version")
    common_worksheet.cell(row_num,5).value=str(received_configs.get("mgmt_addresses"))
    common_worksheet.cell(row_num,6).value=received_configs.get("serial")
    common_worksheet.cell(row_num,7).value=received_configs.get("model")

def update_dns_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):
    ntp_worksheet = received_workbook[received_worksheet_name]
    row_num = ntp_worksheet.max_row+1
    ntp_worksheet.cell(row_num,1).value=received_ar_ip_addr 
    ntp_worksheet.cell(row_num,2).value=received_configs.get("hostname")
    ntp_worksheet.cell(row_num,3).value=str(received_configs.get("dns_servers"))
    ntp_worksheet.cell(row_num,4).value=str(received_configs.get("dns_domains"))

def update_hostgroups_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):
    hg_worksheet = received_workbook[received_worksheet_name]
    row_num = hg_worksheet.max_row+1 
    
    for item in received_configs['items']:
        hg_worksheet.cell(row_num,1).value=received_ar_ip_addr 
        hg_worksheet.cell(row_num,2).value=item.get("name")
        hg_worksheet.cell(row_num,3).value=item.get("desc")
        hg_worksheet.cell(row_num,4).value=item.get("enabled")
        hg_worksheet.cell(row_num,5).value=str(item.get("hosts"))
        hg_worksheet.cell(row_num,6).value=item.get("id")
        hg_worksheet.cell(row_num,7).value=str(item.get("member_hostgroups"))
        hg_worksheet.cell(row_num,8).value=item.get("in_speed")
        hg_worksheet.cell(row_num,9).value=item.get("in_speed_unit")
        hg_worksheet.cell(row_num,10).value=item.get("out_speed")
        hg_worksheet.cell(row_num,11).value=item.get("out_speed_unit")

        row_num += 1

def update_urls_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):
    urls_worksheet = received_workbook[received_worksheet_name]
    row_num = urls_worksheet.max_row+1

    for item in received_configs['items']:
        urls_worksheet.cell(row_num,1).value=received_ar_ip_addr 
        urls_worksheet.cell(row_num,2).value=item.get("name")
        urls_worksheet.cell(row_num,3).value=item.get("desc")
        urls_worksheet.cell(row_num,4).value=item.get("enabled")
        urls_worksheet.cell(row_num,5).value=item.get("preferred")
        urls_worksheet.cell(row_num,6).value=str(item.get("urls"))
        urls_worksheet.cell(row_num,7).value=item.get("id")

def update_apps_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):
    apps_worksheet = received_workbook[received_worksheet_name]
    row_num = apps_worksheet.max_row+1

    for item in received_configs['items']:
        apps_worksheet.cell(row_num,1).value=received_ar_ip_addr 
        apps_worksheet.cell(row_num,2).value=item.get("name")
        apps_worksheet.cell(row_num,10).value=item.get("id")
        apps_worksheet.cell(row_num,3).value=item.get("desc")
        apps_worksheet.cell(row_num,4).value=item.get("enabled")
        apps_worksheet.cell(row_num,5).value=item.get("traffic_match_mode")
        apps_worksheet.cell(row_num,6).value=item.get("include_dpi_tags")
        definitions = item['definitions']
        for def_item in definitions['items']:
            if def_item.get("hosts"):
                apps_worksheet.cell(row_num,7).value=str(def_item.get("hosts"))
            else:
                apps_worksheet.cell(row_num,7).value="[]"
            for trans_rules in def_item['transport_rules']:
                apps_worksheet.cell(row_num,8).value=trans_rules.get("ports")
                apps_worksheet.cell(row_num,9).value=trans_rules.get("ip_protocol")
                row_num += 1

        row_num = apps_worksheet.max_row+1

def update_phy_int_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):
    phy_int_worksheet = received_workbook[received_worksheet_name]
    row_num = phy_int_worksheet.max_row+1
    for item in received_configs["items"]:
        config = item["config"]
        phy_int_worksheet.cell(row_num,1).value=received_ar_ip_addr 
        phy_int_worksheet.cell(row_num,2).value=config.get("alias")
        phy_int_worksheet.cell(row_num,3).value=config.get("enabled")
        phy_int_worksheet.cell(row_num,4).value=config.get("description")
        phy_int_worksheet.cell(row_num,5).value=config.get("speed_duplex")
        state = item["state"]
        phy_int_worksheet.cell(row_num,6).value=state.get("interface_type")
        phy_int_worksheet.cell(row_num,7).value=state.get("status")
        phy_int_worksheet.cell(row_num,8).value=state.get("mtu")

        row_num += 1

def update_cap_jobs_worksheet(received_ar_ip_addr,received_configs,received_workbook,received_worksheet_name):
    cap_jobs_worksheet = received_workbook[received_worksheet_name]
    row_num = cap_jobs_worksheet.max_row+1
    for item in received_configs["items"]:
        config = item["config"]
        cap_jobs_worksheet.cell(row_num,1).value=received_ar_ip_addr 
        cap_jobs_worksheet.cell(row_num,2).value=config.get("name")
        cap_jobs_worksheet.cell(row_num,3).value=str(config.get("vifgs"))
        cap_jobs_worksheet.cell(row_num,4).value=config.get("capture_from_all_vifgs")
        cap_jobs_worksheet.cell(row_num,5).value=config.get("snap_len")
        cap_jobs_worksheet.cell(row_num,6).value=config.get("enabled")
        index = config['indexing']

        cap_jobs_worksheet.cell(row_num,7).value=index.get("enabled")
        retention_rules = index['retention_rules']
        cap_jobs_worksheet.cell(row_num,8).value=retention_rules.get("min_disk_space")
        cap_jobs_worksheet.cell(row_num,9).value=retention_rules.get("max_disk_space")
        cap_jobs_worksheet.cell(row_num,10).value=retention_rules.get("min_retention_time")
        cap_jobs_worksheet.cell(row_num,11).value=retention_rules.get("max_retention_time")
        packet_data_retention_rules = config["retention_rules"]
        cap_jobs_worksheet.cell(row_num,12).value=config.get("optimize_for_read")
        cap_jobs_worksheet.cell(row_num,13).value=packet_data_retention_rules.get("min_disk_space")
        cap_jobs_worksheet.cell(row_num,14).value=packet_data_retention_rules.get("max_disk_space")
        cap_jobs_worksheet.cell(row_num,15).value=packet_data_retention_rules.get("min_retention_time")
        cap_jobs_worksheet.cell(row_num,16).value=packet_data_retention_rules.get("max_retention_time")

        row_num += 1

def get_AR_configuration_and_update_AR_mgmt_spreadsheet(received_username,received_password,received_csv_file_path):

    orginal_ar_list = get_ar_list_from_ar_list_csv_file(received_csv_file_path)

    #if list isn't empty
    if orginal_ar_list:

        active_ar_list=confirm_rest_api_access_to_ARs_and_get_bearer_token(received_username,received_password,orginal_ar_list)

        #will setup spreadsheet and all relevant worksheets
        setup_management_spreadsheet(MGMT_SPREADSHEET_PATH)

        ars = []
        for ar_info in active_ar_list:

            Ar_Device = appresponse_device.AppResponse(ar_info['ar_ip_addr'],ar_info['bear_token'])
 
            update_spreadsheet_with_config(MGMT_SPREADSHEET_PATH,Ar_Device)

            ars.append(Ar_Device)
    else:
        print(f"unable to list of AppResponses from -----> {received_csv_file_path}")

    return ars

def get_ar_list_from_ar_list_csv_file(received_csv_file_path):
    ar_list = []

    try:
        with open(received_csv_file_path,'r') as csv_file:
            #read data in from csv into a dictionary. Import that header information has a header called "ar_list"
            csv_reader = csv.DictReader(csv_file)

            for line in csv_reader:
                ar_list.append(line.get('ar_list'))
    except:
        print(f"faile to find---> {AR_LIST_CSV_FILE_PATH}")

    #if returned import is "[None]" then we convert it to an empty list
    if ar_list == [None]:
        ar_list = []

    return ar_list

def get_snmp_config_from_spreadsheet_for_this_ar(received_config_spreadsheet,received_ar_ip_addr):
    snmp_worksheet = received_config_spreadsheet[SNMP_WORKSHEET_NAME]
    ip_address_column_num = find_column_with_title(AR_LIST_COL_NAME,snmp_worksheet)

    #find row with ar ip address
    for row_num in range(2,snmp_worksheet.max_row+1):
        if received_ar_ip_addr == snmp_worksheet.cell(row_num,ip_address_column_num).value:
            ar_row_num = row_num
            
    snmp_config={}

    snmp_config['enabled']=snmp_worksheet.cell(ar_row_num,2).value
    snmp_config['contact']=snmp_worksheet.cell(ar_row_num,3).value
    snmp_config['location']=snmp_worksheet.cell(ar_row_num,4).value
    snmp_config['description']=snmp_worksheet.cell(ar_row_num,5).value
    
    version_configuration={}
    
    version_configuration['version']=snmp_worksheet.cell(row_num,6).value
    version_configuration['community_string']=snmp_worksheet.cell(row_num,7).value   
    version_configuration['username']=snmp_worksheet.cell(row_num,8).value   
    version_configuration['security_model']=snmp_worksheet.cell(row_num,9).value   
    version_configuration['auth_protocol']=snmp_worksheet.cell(row_num,10).value
    version_configuration['authentication_passphrase']=snmp_worksheet.cell(row_num,11).value  
    version_configuration['privacy_protocol']=snmp_worksheet.cell(row_num,12).value
    version_configuration['privacy_passphrase']=snmp_worksheet.cell(row_num,13).value  

    snmp_config['version_configuration'] = version_configuration

    return snmp_config

def get_dns_config_from_spreadsheet_for_this_ar(received_config_spreadsheet,received_ar_ip_addr):
    dns_worksheet = received_config_spreadsheet[DNS_WORKSHEET_NAME]
    ip_address_column_num = find_column_with_title(AR_LIST_COL_NAME,dns_worksheet)


    for row_num in range(2,dns_worksheet.max_row+1):
        if received_ar_ip_addr == dns_worksheet.cell(row_num,ip_address_column_num).value:
            ar_row_num = row_num

    dns_config={}

    dns_config["hostname"]= dns_worksheet.cell(ar_row_num,2).value 
    dns_config["dns_servers"]= string_to_list(dns_worksheet.cell(ar_row_num,3).value)
    dns_config["dns_domains"]= string_to_list(dns_worksheet.cell(ar_row_num,4).value)

    return dns_config

def get_ntp_config_from_spreadsheet_for_this_ar(received_config_spreadsheet,received_ar_ip_addr):
    ntp_worksheet = received_config_spreadsheet[NTP_WORKSHEET_NAME]
    ip_address_column_num = find_column_with_title(AR_LIST_COL_NAME,ntp_worksheet)
    ntp_config_list={}
    ntp_config_list["items"]=[]
    for row_num in range(2,ntp_worksheet.max_row+1):
        if received_ar_ip_addr == ntp_worksheet.cell(row_num,ip_address_column_num).value:
            ntp_config={}
            ntp_config['server_id']=ntp_worksheet.cell(row_num,2).value
            ntp_config['address']=ntp_worksheet.cell(row_num,3).value
            ntp_config['prefer']=ntp_worksheet.cell(row_num,4).value
            ntp_config['version']=ntp_worksheet.cell(row_num,5).value
            ntp_config['encryption']=ntp_worksheet.cell(row_num,6).value
            ntp_config['key_id']=ntp_worksheet.cell(row_num,7).value
            ntp_config['secret']=ntp_worksheet.cell(row_num,8).value

            ntp_config_list["items"].append(ntp_config)
    return ntp_config_list

def get_vifgs_config_from_spreadsheet_for_this_ar(received_config_spreadsheet,received_ar_ip_addr):
    vifgs_worksheet = received_config_spreadsheet[VIFGS_WORKSHEET_NAME]
    ip_address_column_num = find_column_with_title(AR_LIST_COL_NAME,vifgs_worksheet)
    vifgs_config_list={}
    vifgs_config_list["items"]=[]

    for row_num in range(2,vifgs_worksheet.max_row+1):
        if received_ar_ip_addr == vifgs_worksheet.cell(row_num,ip_address_column_num).value:
            vifgs_config={}
            vifgs_data={}

            vifgs_config["name"]=vifgs_worksheet.cell(row_num,2).value
            vifgs_data["id"]=vifgs_worksheet.cell(row_num,3).value
            vifgs_config["enabled"]=vifgs_worksheet.cell(row_num,4).value
            vifgs_config["dedup"]=vifgs_worksheet.cell(row_num,5).value
            vifgs_config["members"]= string_to_list(vifgs_worksheet.cell(row_num,6).value)
            vifgs_config["description"]=vifgs_worksheet.cell(row_num,7).value
            filter = {}
            filter["type"]=vifgs_worksheet.cell(row_num,8).value
            filter["value"]=vifgs_worksheet.cell(row_num,9).value
            vifgs_config["filter"]=filter
            vifgs_config["bandwidth_capacity"]=vifgs_worksheet.cell(row_num,10).value
            vifgs_config["is_other_vifg"]=vifgs_worksheet.cell(row_num,11).value
            vifgs_data["config"]=vifgs_config
            vifgs_config_list["items"].append(vifgs_data)

    return vifgs_config_list

def get_hostgroups_config_from_spreadsheet_for_this_ar(received_config_spreadsheet,received_ar_ip_addr):
    hostgroups_worksheet = received_config_spreadsheet[HOSTGROUPS_WORKSHEET_NAME]
    ip_address_column_num = find_column_with_title(AR_LIST_COL_NAME,hostgroups_worksheet)
    hostgroups_config_list={}
    hostgroups_config_list["items"]=[]
    for row_num in range(2,hostgroups_worksheet.max_row+1):
        if received_ar_ip_addr == hostgroups_worksheet.cell(row_num,ip_address_column_num).value:
            hostgroup_config={}
            hostgroup_config["name"]=hostgroups_worksheet.cell(row_num,2).value
            hostgroup_config["desc"]=hostgroups_worksheet.cell(row_num,3).value
            hostgroup_config["enabled"]=hostgroups_worksheet.cell(row_num,4).value

            print(f"\n\t\t\t\t\t\tcell data {string_to_list(hostgroups_worksheet.cell(row_num,5).value)}\n")
            hostgroup_config["hosts"]=string_to_list(hostgroups_worksheet.cell(row_num,5).value)
            print()
            if not hostgroups_worksheet.cell(row_num,6).value == None: #no hostgroup id means a new hostgroup. used for telling the difference between updating a current hostgroup or creating a new one
                hostgroup_config["id"]=hostgroups_worksheet.cell(row_num,6).value

            #print(f"\tmemememememememe  \t\t\t\t\t\t{(hostgroups_worksheet.cell(row_num,7).value)} data type is {type((hostgroups_worksheet.cell(row_num,7).value))} length is {len((hostgroups_worksheet.cell(row_num,7).value))}\t {(hostgroups_worksheet.cell(row_num,7).value)[0]}")
            if not hostgroups_worksheet.cell(row_num,7).value == None:
                hostgroup_config["member_hostgroups"]=string_to_list(hostgroups_worksheet.cell(row_num,7).value)

            if not (hostgroups_worksheet.cell(row_num,8).value) == None:
                hostgroup_config["in_speed"]=hostgroups_worksheet.cell(row_num,8).value
            if not (hostgroups_worksheet.cell(row_num,9).value) == None:
                hostgroup_config["in_speed_unit"]=hostgroups_worksheet.cell(row_num,9).value
            if not (hostgroups_worksheet.cell(row_num,10).value) == None:
                hostgroup_config["out_speed"]=hostgroups_worksheet.cell(row_num,10).value
            if not (hostgroups_worksheet.cell(row_num,11).value) == None:    
                hostgroup_config["out_speed_unit"]=hostgroups_worksheet.cell(row_num,11).value

            hostgroups_config_list["items"].append(hostgroup_config)
    return hostgroups_config_list

def get_urls_config_from_spreadsheet_for_this_ar(received_config_spreadsheet,received_ar_ip_addr):
    urls_worksheet = received_config_spreadsheet[URL_WORKSHEET_NAME]
    ip_address_column_num = find_column_with_title(AR_LIST_COL_NAME,urls_worksheet)
    urls_config_list={}
    urls_config_list["items"]=[]
    for row_num in range(2,urls_worksheet.max_row+1):
        if received_ar_ip_addr == urls_worksheet.cell(row_num,ip_address_column_num).value:
            url_config={}
            url_config["name"]=urls_worksheet.cell(row_num,2).value
            url_config["desc"]=urls_worksheet.cell(row_num,3).value
            url_config["enabled"]=urls_worksheet.cell(row_num,4).value
            url_config["preferred"]=urls_worksheet.cell(row_num,5).value 
            url_config["urls"]=string_of_strings_to_list(urls_worksheet.cell(row_num,6).value)
            url_config["id"]=urls_worksheet.cell(row_num,7).value

            urls_config_list["items"].append(url_config)

    return urls_config_list

def get_cap_jobs_config_from_spreadsheet_for_this_ar(received_config_spreadsheet,received_ar_ip_addr):

    cap_jobs_worksheet = received_config_spreadsheet[CAP_JOBS_WORKSHEET_NAME]
    cap_jobs_config_list={}
    cap_jobs_config_list["items"]=[]
    ip_address_column_num = find_column_with_title(AR_LIST_COL_NAME,cap_jobs_worksheet)#
    for row_num in range(2,cap_jobs_worksheet.max_row+1):
        if received_ar_ip_addr == cap_jobs_worksheet.cell(row_num,ip_address_column_num).value:
            cap_job_config={}
            
            config={}
            config["name"]=cap_jobs_worksheet.cell(row_num,2).value
            config["vifgs"]=string_to_list(cap_jobs_worksheet.cell(row_num,3).value)
            config["capture_from_all_vifgs"]=cap_jobs_worksheet.cell(row_num,4).value
            config["snap_len"]=cap_jobs_worksheet.cell(row_num,5).value
            config["enabled"]=cap_jobs_worksheet.cell(row_num,6).value

            index={}
            index["enabled"]=cap_jobs_worksheet.cell(row_num,7).value

            retention_rules={}
            retention_rules["min_disk_space"]=cap_jobs_worksheet.cell(row_num,8).value 
            retention_rules["max_disk_space"]=cap_jobs_worksheet.cell(row_num,9).value
            retention_rules["min_retention_time"]=cap_jobs_worksheet.cell(row_num,10).value
            retention_rules["max_retention_time"]=cap_jobs_worksheet.cell(row_num,11).value

            packet_data_retention_rules = {}
            config["optimize_for_read"]=cap_jobs_worksheet.cell(row_num,12).value
            packet_data_retention_rules["min_disk_space"]=cap_jobs_worksheet.cell(row_num,13).value
            packet_data_retention_rules["max_disk_space"]=cap_jobs_worksheet.cell(row_num,14).value
            packet_data_retention_rules["min_retention_time"]=cap_jobs_worksheet.cell(row_num,15).value
            packet_data_retention_rules["max_retention_time"]=cap_jobs_worksheet.cell(row_num,16).value

            config["retention_rules"]=packet_data_retention_rules
            index['retention_rules']=retention_rules
            config['indexing']=index
            cap_job_config["config"]=config
            cap_jobs_config_list["items"].append(cap_job_config)

    return cap_jobs_config_list

def get_definitions_list_first_and_last_row_number(received_ar_ip_addr,received_worksheet,received_current_row):
    first_row_num_definitions_list = received_current_row
    last_row_num_definitions_list = 0
    for row_num in range(received_current_row+1,received_worksheet.max_row+1):
        if received_worksheet.cell(row_num,1).value != None and received_worksheet.cell(row_num,1).value!= received_ar_ip_addr:
            last_row_num_definitions_list=row_num-1
            break
        elif row_num == received_worksheet.max_row:
            last_row_num_definitions_list=row_num

    print(f"first definition row is {first_row_num_definitions_list}")
    print(f"last definition row is {last_row_num_definitions_list}")
    return(first_row_num_definitions_list,last_row_num_definitions_list)

def get_transport_rules_list_first_and_last_row_number(received_ar_ip_addr,received_worksheet,received_current_row,received_definitions_rules_list_last_row_num):
    print(f"transport received row {received_current_row}")
    first_row_num_transport_rule_list = received_current_row
    last_row_num_transport_rule_list = 0
    for row_num in range(received_current_row,received_definitions_rules_list_last_row_num+1):
        print (f"\t\t{received_worksheet.cell(row_num,1).value} row number is {row_num}")
        if received_worksheet.cell(row_num+1,1).value != None or row_num==received_definitions_rules_list_last_row_num:
            last_row_num_transport_rule_list = row_num
            break
    print(f"first transport rule row is {first_row_num_transport_rule_list}")
    print(f"last transport rule row is {last_row_num_transport_rule_list}")
    return (first_row_num_transport_rule_list,last_row_num_transport_rule_list)

def find_how_many_rows_are_in_an_application_rule(received_worksheet,received_conffig_row):
    first_row_of_app_rule = received_conffig_row
    
    if received_conffig_row < received_worksheet.max_row:
        last_row_of_app_rule = received_conffig_row+1
    else:
        last_row_of_app_rule = received_conffig_row

    if received_worksheet.cell(last_row_of_app_rule,1).value != None: #a one line rule
        last_row_of_app_rule = received_conffig_row

    else: #more than one line application rule

        while received_worksheet.cell(last_row_of_app_rule,1).value == None and last_row_of_app_rule < received_worksheet.max_row:
            last_row_of_app_rule+=1
        if last_row_of_app_rule < received_worksheet.max_row:
            last_row_of_app_rule-=1

    return (first_row_of_app_rule, last_row_of_app_rule)

def find_how_many_definitions_there_are_in_app_rule_and_rows_in_definition(received_worksheet,first_row_of_app_rule,last_row_of_app_rule):
    rule_list = []
    if first_row_of_app_rule == last_row_of_app_rule:

        def_rule_info={"first_row":first_row_of_app_rule,"last_row":last_row_of_app_rule}
        rule_list.append(def_rule_info)
        return rule_list
    else:
        def_rule_count = 0
        for row_num in range(first_row_of_app_rule,last_row_of_app_rule+1):

            def_rule_count+=1

            if received_worksheet.cell(row_num,7).value != None:
                def_rule_count = 1

            
            if def_rule_count == 1 and row_num == first_row_of_app_rule:
                def_rule_info={}
                def_rule_info["first_row"]=row_num
                #print("this is the first rule")

            elif def_rule_count == 1 and row_num == last_row_of_app_rule:
         
                def_rule_info["last_row"]=row_num-1
                rule_list.append(def_rule_info)
                def_rule_info={}
                def_rule_info["first_row"]=row_num
                def_rule_info["last_row"]=row_num
                rule_list.append(def_rule_info)

            elif def_rule_count == 1:
                def_rule_info["last_row"]=row_num-1
                rule_list.append(def_rule_info)
                def_rule_info={}
                def_rule_info["first_row"]=row_num
                #print("this is the next rule")
            elif row_num == last_row_of_app_rule:
                rule_list.append(def_rule_info)
                def_rule_info["last_row"]=row_num
                #print("this is the last rule")

        return rule_list

def get_definition_data_from_spreadsheet(received_worksheet,received_app_rule_rows_list):
    definition_items=[]
    for app_rule in received_app_rule_rows_list:
        definition_item={}
        transport_rule_list=[]
        for row_num in range(app_rule["first_row"],app_rule["last_row"]+1):
            transport_rule = {}
            if row_num == app_rule["first_row"]:

                if received_worksheet.cell(row_num,7).value != "None" or received_worksheet.cell(row_num,7).value == None:
                    definition_item["hosts"]=string_to_list(received_worksheet.cell(row_num,7).value)


            if received_worksheet.cell(row_num,8).value != None:
                transport_rule['ports'] = received_worksheet.cell(row_num,8).value 
                #print(f"cell vlaue is {received_worksheet.cell(row_num,7).value}, data type is {type(received_worksheet.cell(row_num,7).value)}")
            if received_worksheet.cell(row_num,9).value != None:
                transport_rule['ip_protocol'] = received_worksheet.cell(row_num,9).value 
                #print(received_worksheet.cell(row_num,8).value)
            
            transport_rule_list.append(transport_rule)
        definition_item['transport_rules']=transport_rule_list
        definition_items.append(definition_item)

    return definition_items         
  
def get_app_rules_definitions(received_worksheet,received_conffig_row):
    print(f"\there at the def rules, config row number is {received_conffig_row}")
    
    definition_items_list=[]
    first_row_of_app_rule, last_row_of_app_rule = find_how_many_rows_are_in_an_application_rule(received_worksheet,received_conffig_row)
    print(f"first row of app is:- {first_row_of_app_rule}\t last row is {last_row_of_app_rule}")
    print(f"max row number{received_worksheet.max_row}")
    app_rule_rows_list = find_how_many_definitions_there_are_in_app_rule_and_rows_in_definition(received_worksheet,first_row_of_app_rule,last_row_of_app_rule)
    #print(f"\t\t\there at the def rules, app rule row number is {app_rule_rows_list}")
    definition_items_list = get_definition_data_from_spreadsheet(received_worksheet,app_rule_rows_list)
    return definition_items_list
    
def get_application_config_from_spreadsheet_for_this_ar(received_config_spreadsheet,received_ar_ip_addr):
    apps_worksheet = received_config_spreadsheet[APPLICATIONS_WORKSHEET_NAME]
    ip_address_column_num = find_column_with_title(AR_LIST_COL_NAME,apps_worksheet)
    apps_config_list={}
    apps_config_list["items"]=[]

    #find rows that have the appresponse ip address
    app_config_row_list=[]
    for row_num in range(2,apps_worksheet.max_row+1):
        if apps_worksheet.cell(row_num,ip_address_column_num).value == received_ar_ip_addr:
            app_config_row_list.append(row_num)
    for conffig_row in app_config_row_list:
        item={}

        item["name"]=apps_worksheet.cell(conffig_row,2).value
        item["id"]=apps_worksheet.cell(conffig_row,10).value
        item["desc"]=apps_worksheet.cell(conffig_row,3).value
        item["enabled"]=apps_worksheet.cell(conffig_row,4).value
        item["traffic_match_mode"]=apps_worksheet.cell(conffig_row,5).value
        item["include_dpi_tags"]=apps_worksheet.cell(conffig_row,6).value
        
        definitions={}
        item["definitions"]=definitions
        definitions["items"]=[]
        print(f"\n\n\nearlier max row number {apps_worksheet.max_row}")
        definitions["items"]=get_app_rules_definitions(apps_worksheet,conffig_row)

        apps_config_list["items"].append(item)

    return apps_config_list

def get_config_of_what_user_wants_configured_for_this_ar_from_spreadsheet(received_ar_ip_addr,received_spreadsheet_path,received_what_to_configure):
    #open spreadsheet
    ar_config_spreadsheet = openpyxl.load_workbook(received_spreadsheet_path)
    return_configs_for_ar={}
    
    #go through list of what to configure
  
    if received_what_to_configure == "snmp":
        snmp_config = get_snmp_config_from_spreadsheet_for_this_ar(ar_config_spreadsheet,received_ar_ip_addr)
        return_configs_for_ar["snmp"] = snmp_config
    elif received_what_to_configure == "dns":
        dns_config = get_dns_config_from_spreadsheet_for_this_ar(ar_config_spreadsheet,received_ar_ip_addr)
        return_configs_for_ar["dns"] = dns_config
    elif received_what_to_configure == "ntp":
        ntp_config = get_ntp_config_from_spreadsheet_for_this_ar(ar_config_spreadsheet,received_ar_ip_addr)
        return_configs_for_ar["ntp"] = ntp_config
    #elif received_what_to_configure == "vifgs":
    #    vifgs_config = get_vifgs_config_from_spreadsheet_for_this_ar(ar_config_spreadsheet,received_ar_ip_addr)
    #    return_configs_for_ar["vifgs"] = vifgs_config
    elif received_what_to_configure == "hostgroups":
        hostgroups_config = get_hostgroups_config_from_spreadsheet_for_this_ar(ar_config_spreadsheet,received_ar_ip_addr)
        return_configs_for_ar["hostgroups"] = hostgroups_config
    elif received_what_to_configure == "urls":
        urls_config = get_urls_config_from_spreadsheet_for_this_ar(ar_config_spreadsheet,received_ar_ip_addr)
        return_configs_for_ar["urls"] = urls_config
    #elif received_what_to_configure == "cap_jobs":
    #    cap_jobs_config = get_cap_jobs_config_from_spreadsheet_for_this_ar(ar_config_spreadsheet,received_ar_ip_addr)
    #    return_configs_for_ar["cap_jobs"] = cap_jobs_config
    elif received_what_to_configure == "apps":
        apps_jobs_config = get_application_config_from_spreadsheet_for_this_ar(ar_config_spreadsheet,received_ar_ip_addr)
        return_configs_for_ar["apps"] = apps_jobs_config

    return return_configs_for_ar

def get_ar_configuration_from_spreadsheet_and_push_out_ars(received_list_of_ars,what_to_configure):
    print("=========================================================")
    for ar_device in received_list_of_ars:
        print(ar_device.ip_addr)
        new_ar_config = get_config_of_what_user_wants_configured_for_this_ar_from_spreadsheet(ar_device.ip_addr,MGMT_SPREADSHEET_PATH,what_to_configure)

        if ar_device.affirm_new_config_is_good(new_ar_config):
            ar_device.print_old_and_new_config(new_ar_config)
   
def check_new_configs_are_correct_for_config_option(received_list_of_ars,received_spreadsheet_path,received_config_option):

    configs_are_good = True

    for ar_device in received_list_of_ars:
        new_ar_config = get_config_of_what_user_wants_configured_for_this_ar_from_spreadsheet(ar_device.ip_addr,received_spreadsheet_path,received_config_option)

        if not ar_device.affirm_new_config_is_good(new_ar_config):
            configs_are_good = False
    
    return configs_are_good

def print_out_old_and_new_config(received_list_of_ars,received_config_option):
    for ar_device in received_list_of_ars:
        ar_device.print_old_and_new_config_plus_config_status(received_config_option)

def update_ars_with_new_config(received_list_of_ars,received_config_option):
    for ar_device in received_list_of_ars:
        ar_device.update_ar_device_config(received_config_option)


def  main(received_username,received_password):
    pass

if __name__ == "__main__":
    date_time = datetime.now()
    str_date_time = date_time.strftime("%d-%m-%H-%M-%S")
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-u","--username", required=True,help="username")
    args = vars(ap.parse_args())
    username = format(args["username"])
    password = getpass()
    list_of_ars = get_AR_configuration_and_update_AR_mgmt_spreadsheet(username,password,AR_LIST_CSV_FILE_PATH)

    ar_config_work_book = openpyxl.load_workbook(MGMT_SPREADSHEET_PATH)
    backup_ar_mgmt_spreadsheet_name = f"backup-{str_date_time}.xlsx"
    ar_config_work_book.save(backup_ar_mgmt_spreadsheet_name)

    print(f"\n\nA spreadsheet called {MGMT_SPREADSHEET_PATH} has been created or over written a previous one. With the current config of your AR11's.\n A backup copy has been created '{backup_ar_mgmt_spreadsheet_name}'")

    print("Please reveiw and update the spreadsheet with any configuration changes.")
    print("\n\nOnce you have finished updating the spreadsheet,")
    print(f"Category options are {CONFIG_OPTIONS_LIST}")
    config_options = input("To exit, type 'exit' or leave blank.\n>>>>> ")
    
    while not config_options in ["exit",""]:

        if config_options in CONFIG_OPTIONS_LIST:

            all_new_configs_are_good = check_new_configs_are_correct_for_config_option(list_of_ars,MGMT_SPREADSHEET_PATH,config_options)

            print_out_old_and_new_config(list_of_ars,config_options)
            
            if all_new_configs_are_good:
                continue_options = ""
                user_wants_to_push_out_config = True
                while not continue_options in ["y","n"]:
                    print("\n\nPlease confirm you wish to push out new config: 'y' for yes or 'n' for no")                
                    continue_options = input(">>>>> ")

                if continue_options == 'y':
                   update_ars_with_new_config(list_of_ars,config_options)


            print(f"Config options are {CONFIG_OPTIONS_LIST}")
            config_options = input("To exit, type 'ex""it' or leave blank.\n>>>>> ")

    """
            for ar_device in list_of_ars:
                new_ar_config = get_config_of_what_user_wants_configured_for_this_ar_from_spreadsheet(ar_device.ip_addr,MGMT_SPREADSHEET_PATH,config_options)



            if config_options in CONFIG_OPTIONS_LIST:
                print("\n###############\t##############\t#############\n\tgetting new config\n###############\t##############\t#############\n")
                get_ar_configuration_from_spreadsheet_and_push_out_ars(list_of_ars,config_options)
    """
            
